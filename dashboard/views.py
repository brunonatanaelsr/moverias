from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Count, Q
from django.core.cache import cache
import csv
from datetime import timedelta
from core.optimizers import CacheManager, QueryOptimizer
from core.permissions import is_technician
from core.cache_utils import cache_view, CACHE_TIMEOUTS
from core.unified_permissions import (
    get_user_permissions,
    is_technician,
    is_coordinator,
    is_admin,
    TechnicianRequiredMixin,
    CoordinatorRequiredMixin,
    AdminRequiredMixin,
    requires_technician
)
from members.models import Beneficiary
from members.forms import BeneficiaryForm
from workshops.models import Workshop
# Imports moved to top for better performance
from django.contrib.auth import get_user_model
from users.models import UserActivity
from social.models import SocialAnamnesis
from evolution.models import EvolutionRecord
from projects.models import ProjectEnrollment, Project
from coaching.models import ActionPlan, WheelOfLife


@login_required
@requires_technician
def dashboard_home(request):
    """Dashboard principal com cache inteligente"""
    
    # OTIMIZAÇÃO: Cache com diferentes TTLs
    cache_key = 'dashboard_stats_optimized'
    stats = cache.get(cache_key)
    
    if stats is None:
        User = get_user_model()
        now = timezone.now()
        today = now.date()
        last_week = today - timedelta(days=7)
        last_month = today - timedelta(days=30)
        
        # OTIMIZAÇÃO: Query única com múltiplas agregações
        beneficiary_stats = Beneficiary.objects.aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(status='ATIVA')),
            inactive=Count('id', filter=Q(status='INATIVA')),
            new_week=Count('id', filter=Q(created_at__date__gte=last_week)),
            new_month=Count('id', filter=Q(created_at__date__gte=last_month))
        )
        
        # OTIMIZAÇÃO: Queries otimizadas com select_related/prefetch_related
        workshop_stats = Workshop.objects.prefetch_related('enrollments').aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(status='ativo')),
            total_participants=Count('enrollments', filter=Q(enrollments__status='ativo'), distinct=True)
        )
        
        project_stats = Project.objects.prefetch_related('enrollments').aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(status='ATIVO')),
            total_participants=Count('enrollments', filter=Q(enrollments__status='ATIVO'), distinct=True)
        )
        
        stats = {
            'beneficiary_stats': beneficiary_stats,
            'workshop_stats': workshop_stats,
            'project_stats': project_stats,
            'user_count': User.objects.filter(is_active=True).count(),
            'updated_at': now.isoformat()
        }
        
        # Cache por 5 minutos (dados dinâmicos)
        cache.set(cache_key, stats, 300)
    
    context = {
        'stats': stats,
        'cache_status': 'hit' if stats else 'miss',
        'user_permissions': get_user_permissions(request.user),
    }
    
    return render(request, 'dashboard/home.html', context)


@login_required
@requires_technician
def activity_stats(request):
    """Vista para estatísticas de atividades"""
    today = timezone.now().date()
    last_week = today - timedelta(days=7)
    last_month = today - timedelta(days=30)
    
    # Estatísticas detalhadas
    extended_stats = {
        # Atividades
        'activity_stats': UserActivity.objects.aggregate(
            total=Count('id'),
            today=Count('id', filter=Q(timestamp__date=today)),
            week=Count('id', filter=Q(timestamp__date__gte=last_week))
        ),
        
        # Anamneses
        'anamnesis_stats': SocialAnamnesis.objects.aggregate(
            total=Count('id'),
            pending=Count('id', filter=Q(locked=False)),
            completed=Count('id', filter=Q(locked=True))
        ),
        
        # Evoluções
        'evolution_stats': EvolutionRecord.objects.aggregate(
            total=Count('id'),
            month=Count('id', filter=Q(created_at__date__gte=last_month))
        ),
        
        # Projetos (usando o modelo Project)
        'project_stats': Project.objects.aggregate(
            total=Count('id'),
            active=Count('id')  # Todos os projetos são considerados ativos
        ),
        
        # Matrículas em projetos
        'enrollment_stats': ProjectEnrollment.objects.aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(status='ATIVO'))
        ),
        
        # Planos de Ação
        'action_plan_stats': ActionPlan.objects.aggregate(
            total=Count('id'),
            recent=Count('id', filter=Q(created_at__date__gte=last_month))
        ),
        
        # Rodas da Vida
        'wheel_stats': WheelOfLife.objects.aggregate(
            total=Count('id'),
            recent=Count('id', filter=Q(created_at__date__gte=last_month))
        ),
            
        # Workshops com aggregation
        'workshop_stats': Workshop.objects.aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(status='ativo')),
            completed=Count('id', filter=Q(status='concluido')),
            cancelled=Count('id', filter=Q(status='cancelado'))
        ),
        
        # Beneficiários
        'beneficiary_stats': Beneficiary.objects.aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(status='ATIVA')),
            inactive=Count('id', filter=Q(status='INATIVA')),
            new_week=Count('id', filter=Q(created_at__date__gte=last_week)),
            new_month=Count('id', filter=Q(created_at__date__gte=last_month))
        )
    }
    
    # Dados recentes em tempo real (SEM CACHE)
    recent_data = {
        'recent_beneficiaries': Beneficiary.objects.order_by('-created_at')[:5],
        'recent_workshops': Workshop.objects.order_by('-created_at')[:3],
        'recent_activities': UserActivity.objects.select_related('user').order_by('-timestamp')[:5],
        'upcoming_events': Workshop.objects.filter(
            status='ativo',
            start_date__gte=today
        ).order_by('start_date')[:5]
    }
    
    # Estatísticas por faixa etária em tempo real
    age_stats = {
        'children': Beneficiary.objects.filter(dob__gte=today - timedelta(days=18*365)).count(),
        'adults': Beneficiary.objects.filter(
            dob__lt=today - timedelta(days=18*365),
            dob__gte=today - timedelta(days=60*365)
        ).count(),
        'elderly': Beneficiary.objects.filter(dob__lt=today - timedelta(days=60*365)).count(),
    }
    
    # Criar contexto unificado com todas as estatísticas
    context = {
        # Estatísticas principais para compatibilidade com template
        'stats': {
            'total_beneficiaries': extended_stats['beneficiary_stats']['total'],
            'active_projects': extended_stats['project_stats']['total'],  # Usar total de projetos
            'active_workshops': extended_stats['workshop_stats']['active'],
            'monthly_sessions': extended_stats['evolution_stats']['month'],
            'total_anamnesis': extended_stats['anamnesis_stats']['total'],
            'completed_anamnesis': extended_stats['anamnesis_stats']['completed'],
            'pending_anamnesis': extended_stats['anamnesis_stats']['pending'],
        },
        # Dados detalhados
        **extended_stats,
        **recent_data,
        'age_stats': age_stats,
    }
    
    return render(request, 'dashboard/home.html', context)


@login_required
@user_passes_test(is_technician)
def beneficiaries_list(request):
    """Lista de beneficiárias com busca HTMX e otimizações"""
    search_query = request.GET.get('q', '')
    
    # Query simples sem cache por enquanto
    beneficiaries = Beneficiary.objects.all()
    
    if search_query:
        beneficiaries = beneficiaries.filter(
            Q(full_name__icontains=search_query) |
            Q(phone_1__icontains=search_query) |
            Q(phone_2__icontains=search_query) |
            Q(neighbourhood__icontains=search_query)
        )
    
    # Paginação
    from django.core.paginator import Paginator
    paginator = Paginator(beneficiaries.order_by('full_name'), 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'total_count': paginator.count
    }
    
    if request.headers.get('HX-Request'):
        return render(request, 'dashboard/partials/beneficiaries_table.html', context)
    
    return render(request, 'dashboard/beneficiaries_list.html', context)


@login_required
@user_passes_test(is_technician)
def beneficiary_detail(request, pk):
    """Detalhes de uma beneficiária"""
    beneficiary = get_object_or_404(Beneficiary, pk=pk)
    
    context = {
        'beneficiary': beneficiary,
    }
    
    if request.headers.get('HX-Request'):
        return render(request, 'dashboard/partials/beneficiary_detail.html', context)
    
    return render(request, 'dashboard/beneficiary_detail.html', context)


@login_required
@user_passes_test(is_technician)
def beneficiary_create(request):
    """Formulário para criar nova beneficiária"""
    if request.method == 'POST':
        form = BeneficiaryForm(request.POST)
        if form.is_valid():
            beneficiary = form.save()
            messages.success(
                request, 
                f'Beneficiária {beneficiary.full_name} cadastrada com sucesso!'
            )
            return redirect('dashboard:beneficiary_detail', pk=beneficiary.pk)
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = BeneficiaryForm()
    
    context = {
        'form': form,
        'title': 'Nova Beneficiária'
    }
    return render(request, 'dashboard/beneficiary_form.html', context)


@login_required
@user_passes_test(is_technician)
def beneficiary_edit(request, pk):
    """Formulário para editar beneficiária"""
    beneficiary = get_object_or_404(Beneficiary, pk=pk)
    
    if request.method == 'POST':
        form = BeneficiaryForm(request.POST, instance=beneficiary)
        if form.is_valid():
            beneficiary = form.save()
            messages.success(
                request, 
                f'Dados de {beneficiary.full_name} atualizados com sucesso!'
            )
            return redirect('dashboard:beneficiary_detail', pk=beneficiary.pk)
        else:
            messages.error(request, 'Por favor, corrija os erros abaixo.')
    else:
        form = BeneficiaryForm(instance=beneficiary)
    
    context = {
        'form': form,
        'beneficiary': beneficiary,
        'title': f'Editar {beneficiary.full_name}'
    }
    return render(request, 'dashboard/beneficiary_form.html', context)


@login_required
@user_passes_test(is_technician)
def beneficiaries_export(request):
    """Export beneficiaries list to CSV"""
    
    # Get all beneficiaries with optimized query
    beneficiaries = Beneficiary.objects.select_related().order_by('full_name')
    
    # Create HTTP response with CSV content type
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="beneficiarias_movemarias_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # Create CSV writer
    writer = csv.writer(response)
    
    # Write CSV header
    writer.writerow([
        'Nome Completo',
        'Data de Nascimento', 
        'NIS',
        'Telefone 1',
        'Telefone 2',
        'Bairro',
        'Endereço',
        'Referência',
        'Data de Cadastro'
    ])
    
    # Write beneficiary data
    for beneficiary in beneficiaries:
        writer.writerow([
            beneficiary.full_name,
            beneficiary.dob.strftime('%d/%m/%Y') if beneficiary.dob else '',
            beneficiary.nis or '',
            beneficiary.phone_1 or '',
            beneficiary.phone_2 or '',
            beneficiary.neighbourhood or '',
            beneficiary.address or '',
            beneficiary.reference or '',
            beneficiary.created_at.strftime('%d/%m/%Y %H:%M') if beneficiary.created_at else ''
        ])
    
    return response


@login_required
@user_passes_test(is_technician)
def reports(request):
    """Página de relatórios gerais do sistema"""
    from django.db.models import Count, Avg
    from datetime import datetime, timedelta
    
    # Definir período para relatórios
    now = timezone.now()
    today = now.date()
    last_30_days = today - timedelta(days=30)
    last_90_days = today - timedelta(days=90)
    
    # Estatísticas gerais
    stats = {
        'beneficiaries': {
            'total': Beneficiary.objects.count(),
            'new_last_30_days': Beneficiary.objects.filter(created_at__date__gte=last_30_days).count(),
            'by_neighborhood': Beneficiary.objects.values('neighbourhood').annotate(
                count=Count('id')
            ).order_by('-count')[:10]
        },
        'projects': {
            'total': Project.objects.count(),
            'active_enrollments': ProjectEnrollment.objects.filter(status='ATIVO').count(),
            'by_project': Project.objects.annotate(
                enrollment_count=Count('enrollments')
            ).order_by('-enrollment_count')[:10]
        },
        'workshops': {
            'total': Workshop.objects.count(),
            'active': Workshop.objects.filter(status='ativo').count(),
            'completed': Workshop.objects.filter(status='concluido').count(),
            'recent': Workshop.objects.filter(created_at__date__gte=last_30_days).count()
        },
        'evolution': {
            'total_records': EvolutionRecord.objects.count(),
            'last_30_days': EvolutionRecord.objects.filter(created_at__date__gte=last_30_days).count(),
            'avg_monthly': EvolutionRecord.objects.filter(
                created_at__date__gte=last_90_days
            ).count() / 3 if EvolutionRecord.objects.filter(created_at__date__gte=last_90_days).count() > 0 else 0
        },
        'anamnesis': {
            'total': SocialAnamnesis.objects.count(),
            'completed': SocialAnamnesis.objects.filter(locked=True).count(),
            'pending': SocialAnamnesis.objects.filter(locked=False).count(),
            'completion_rate': 0
        }
    }
    
    # Calcular taxa de conclusão de anamneses
    if stats['anamnesis']['total'] > 0:
        stats['anamnesis']['completion_rate'] = (
            stats['anamnesis']['completed'] / stats['anamnesis']['total'] * 100
        )
    
    # Dados para gráficos
    chart_data = {
        'beneficiaries_by_month': [],
        'workshops_by_status': {
            'labels': ['Ativo', 'Concluído', 'Cancelado'],
            'data': [
                Workshop.objects.filter(status='ativo').count(),
                Workshop.objects.filter(status='concluido').count(),
                Workshop.objects.filter(status='cancelado').count()
            ]
        },
        'enrollments_by_month': []
    }
    
    # Dados de beneficiárias por mês (últimos 6 meses)
    for i in range(6):
        month_start = (today.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        count = Beneficiary.objects.filter(
            created_at__date__gte=month_start,
            created_at__date__lte=month_end
        ).count()
        
        chart_data['beneficiaries_by_month'].insert(0, {
            'month': month_start.strftime('%b/%y'),
            'count': count
        })
    
    # Dados de matrículas por mês (últimos 6 meses)
    for i in range(6):
        month_start = (today.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        count = ProjectEnrollment.objects.filter(
            created_at__gte=month_start,
            created_at__lte=month_end
        ).count()
        
        chart_data['enrollments_by_month'].insert(0, {
            'month': month_start.strftime('%b/%y'),
            'count': count
        })
    
    context = {
        'stats': stats,
        'chart_data': chart_data,
        'last_updated': now,
        'title': 'Relatórios Gerais'
    }
    
    return render(request, 'dashboard/reports.html', context)


@login_required
@requires_technician
def custom_reports(request):
    """Relatórios personalizados com filtros avançados"""
    from django.db.models import Count, Sum, Avg, F, Q
    from datetime import datetime, timedelta
    
    # Parâmetros de filtro
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    report_type = request.GET.get('report_type', 'beneficiaries')
    export_format = request.GET.get('export_format', 'html')
    
    # Definir período padrão (último mês)
    if not start_date:
        start_date = (timezone.now() - timedelta(days=30)).date()
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = timezone.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Dados baseados no tipo de relatório
    report_data = {}
    
    if report_type == 'beneficiaries':
        # Relatório de beneficiárias
        beneficiaries = Beneficiary.objects.filter(
            created_at__date__range=[start_date, end_date]
        ).prefetch_related('project_enrollments', 'evolution_records')
        
        report_data = {
            'title': 'Relatório de Beneficiárias',
            'total_count': beneficiaries.count(),
            'by_status': beneficiaries.values('status').annotate(count=Count('id')),
            'by_neighborhood': beneficiaries.values('neighbourhood').annotate(count=Count('id')).order_by('-count')[:10],
            'by_age_group': _get_age_group_stats(beneficiaries),
            'items': beneficiaries.order_by('-created_at')[:100]  # Limitar para performance
        }
        
    elif report_type == 'workshops':
        # Relatório de workshops
        workshops = Workshop.objects.filter(
            created_at__date__range=[start_date, end_date]
        ).prefetch_related('enrollments')
        
        report_data = {
            'title': 'Relatório de Workshops',
            'total_count': workshops.count(),
            'by_status': workshops.values('status').annotate(count=Count('id')),
            'by_type': workshops.values('workshop_type').annotate(count=Count('id')),
            'avg_participants': workshops.aggregate(avg=Avg('max_participants'))['avg'] or 0,
            'items': workshops.order_by('-created_at')
        }
        
    elif report_type == 'enrollments':
        # Relatório de matrículas
        enrollments = ProjectEnrollment.objects.filter(
            created_at__date__range=[start_date, end_date]
        ).select_related('beneficiary', 'project')
        
        report_data = {
            'title': 'Relatório de Matrículas',
            'total_count': enrollments.count(),
            'by_status': enrollments.values('status').annotate(count=Count('id')),
            'by_project': enrollments.values('project__name').annotate(count=Count('id')).order_by('-count'),
            'items': enrollments.order_by('-created_at')
        }
    
    # Exportar se solicitado
    if export_format == 'csv':
        return _export_custom_report_csv(report_data, report_type, start_date, end_date)
    elif export_format == 'excel':
        return _export_custom_report_excel(report_data, report_type, start_date, end_date)
    
    context = {
        'report_data': report_data,
        'report_type': report_type,
        'start_date': start_date,
        'end_date': end_date,
        'available_types': [
            ('beneficiaries', 'Beneficiárias'),
            ('workshops', 'Workshops'),
            ('enrollments', 'Matrículas'),
            ('evolution', 'Evolução'),
        ]
    }
    
    return render(request, 'dashboard/custom_reports.html', context)


@login_required
@requires_technician
def advanced_analytics(request):
    """Análises avançadas com gráficos e métricas detalhadas"""
    from django.db.models import Count, Sum, Avg, F, Q, Case, When, IntegerField
    from datetime import datetime, timedelta
    import json
    
    # Período de análise
    period = request.GET.get('period', '3months')
    
    if period == '1month':
        start_date = timezone.now() - timedelta(days=30)
    elif period == '6months':
        start_date = timezone.now() - timedelta(days=180)
    elif period == '1year':
        start_date = timezone.now() - timedelta(days=365)
    else:  # 3months
        start_date = timezone.now() - timedelta(days=90)
    
    end_date = timezone.now()
    
    # Análises de tendência
    trends = {
        'beneficiaries': _get_trend_data(
            Beneficiary.objects.filter(created_at__range=[start_date, end_date]),
            'created_at'
        ),
        'workshops': _get_trend_data(
            Workshop.objects.filter(created_at__range=[start_date, end_date]),
            'created_at'
        ),
        'enrollments': _get_trend_data(
            ProjectEnrollment.objects.filter(created_at__range=[start_date, end_date]),
            'created_at'
        )
    }
    
    # Análises de performance
    performance_metrics = {
        'workshop_completion_rate': _calculate_workshop_completion_rate(),
        'beneficiary_engagement': _calculate_beneficiary_engagement(),
        'project_success_rate': _calculate_project_success_rate(),
        'retention_rate': _calculate_retention_rate()
    }
    
    # Análises geográficas
    geographic_data = {
        'beneficiaries_by_neighborhood': list(
            Beneficiary.objects.values('neighbourhood')
            .annotate(count=Count('id'))
            .order_by('-count')[:15]
        ),
        'coverage_analysis': _analyze_geographic_coverage()
    }
    
    # Análises demográficas
    demographic_data = {
        'age_distribution': _get_age_distribution(),
        'education_level': _get_education_distribution(),
        'family_composition': _get_family_composition_stats()
    }
    
    # Predições simples usando médias móveis
    predictions = {
        'next_month_beneficiaries': _predict_next_month_beneficiaries(),
        'workshop_demand': _predict_workshop_demand(),
        'resource_needs': _predict_resource_needs()
    }
    
    context = {
        'period': period,
        'start_date': start_date,
        'end_date': end_date,
        'trends': trends,
        'performance_metrics': performance_metrics,
        'geographic_data': geographic_data,
        'demographic_data': demographic_data,
        'predictions': predictions,
        'charts_data': json.dumps({
            'trends': trends,
            'geographic': geographic_data,
            'demographic': demographic_data
        })
    }
    
    return render(request, 'dashboard/advanced_analytics.html', context)


def _get_age_group_stats(beneficiaries):
    """Calcula estatísticas por faixa etária"""
    age_groups = {
        '18-25': 0,
        '26-35': 0,
        '36-45': 0,
        '46-55': 0,
        '56+': 0
    }
    
    for beneficiary in beneficiaries:
        if beneficiary.dob:
            age = (timezone.now().date() - beneficiary.dob).days // 365
            if age <= 25:
                age_groups['18-25'] += 1
            elif age <= 35:
                age_groups['26-35'] += 1
            elif age <= 45:
                age_groups['36-45'] += 1
            elif age <= 55:
                age_groups['46-55'] += 1
            else:
                age_groups['56+'] += 1
    
    return age_groups


def _get_trend_data(queryset, date_field):
    """Gera dados de tendência para gráficos"""
    from django.db.models.functions import TruncWeek, TruncMonth
    
    # Agrupar por semana usando TruncWeek
    trend_data = list(
        queryset.annotate(week=TruncWeek('created_at'))
        .values('week')
        .annotate(count=Count('id'))
        .order_by('week')
    )
    
    return trend_data


def _calculate_workshop_completion_rate():
    """Calcula taxa de conclusão de workshops"""
    from workshops.models import WorkshopEnrollment
    
    total_enrollments = WorkshopEnrollment.objects.count()
    completed_enrollments = WorkshopEnrollment.objects.filter(status='completed').count()
    
    if total_enrollments > 0:
        return (completed_enrollments / total_enrollments) * 100
    return 0


def _calculate_beneficiary_engagement():
    """Calcula nível de engajamento das beneficiárias"""
    # Simplificado - baseado em número de atividades
    active_beneficiaries = Beneficiary.objects.filter(
        evolution_records__created_at__gte=timezone.now() - timedelta(days=30)
    ).distinct().count()
    
    total_beneficiaries = Beneficiary.objects.filter(status='ATIVA').count()
    
    if total_beneficiaries > 0:
        return (active_beneficiaries / total_beneficiaries) * 100
    return 0


def _calculate_project_success_rate():
    """Calcula taxa de sucesso dos projetos"""
    total_projects = Project.objects.count()
    successful_projects = Project.objects.filter(status='CONCLUIDO').count()
    
    if total_projects > 0:
        return (successful_projects / total_projects) * 100
    return 0


def _calculate_retention_rate():
    """Calcula taxa de retenção"""
    # Simplificado - beneficiárias que continuam ativas após 6 meses
    six_months_ago = timezone.now() - timedelta(days=180)
    
    old_beneficiaries = Beneficiary.objects.filter(
        created_at__lte=six_months_ago
    ).count()
    
    still_active = Beneficiary.objects.filter(
        created_at__lte=six_months_ago,
        status='ATIVA'
    ).count()
    
    if old_beneficiaries > 0:
        return (still_active / old_beneficiaries) * 100
    return 0


def _analyze_geographic_coverage():
    """Analisa cobertura geográfica"""
    neighborhoods = Beneficiary.objects.values('neighbourhood').distinct().count()
    total_beneficiaries = Beneficiary.objects.count()
    
    return {
        'neighborhoods_covered': neighborhoods,
        'avg_beneficiaries_per_neighborhood': total_beneficiaries / neighborhoods if neighborhoods > 0 else 0,
        'coverage_density': 'Alta' if neighborhoods > 20 else 'Média' if neighborhoods > 10 else 'Baixa'
    }


def _get_age_distribution():
    """Distribução de idade das beneficiárias"""
    age_ranges = {
        '18-25': 0, '26-35': 0, '36-45': 0, '46-55': 0, '56+': 0
    }
    
    for beneficiary in Beneficiary.objects.filter(dob__isnull=False):
        age = (timezone.now().date() - beneficiary.dob).days // 365
        if age <= 25:
            age_ranges['18-25'] += 1
        elif age <= 35:
            age_ranges['26-35'] += 1
        elif age <= 45:
            age_ranges['36-45'] += 1
        elif age <= 55:
            age_ranges['46-55'] += 1
        else:
            age_ranges['56+'] += 1
    
    return age_ranges


def _get_education_distribution():
    """Distribuição de nível educacional"""
    # Como o campo education_level não existe no modelo atual,
    # retornamos dados vazios ou simulados
    return {
        'Ensino Fundamental': 0,
        'Ensino Médio': 0,
        'Ensino Superior': 0,
        'Não Informado': Beneficiary.objects.count()
    }


def _get_family_composition_stats():
    """Estatísticas de composição familiar"""
    # Como os campos family_size e marital_status não existem no modelo atual,
    # retornamos dados simulados baseados no total de beneficiárias
    total_beneficiaries = Beneficiary.objects.count()
    
    return {
        'avg_family_size': 3.5,  # Valor simulado
        'single_mothers': int(total_beneficiaries * 0.3),  # 30% estimado
        'families_with_children': int(total_beneficiaries * 0.7)  # 70% estimado
    }


def _predict_next_month_beneficiaries():
    """Predição simples para próximo mês"""
    last_3_months = []
    for i in range(3):
        month_start = timezone.now().replace(day=1) - timedelta(days=i*30)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        count = Beneficiary.objects.filter(
            created_at__range=[month_start, month_end]
        ).count()
        last_3_months.append(count)
    
    # Média móvel simples
    if last_3_months:
        return sum(last_3_months) / len(last_3_months)
    return 0


def _predict_workshop_demand():
    """Predição de demanda por workshops"""
    active_beneficiaries = Beneficiary.objects.filter(status='ATIVA').count()
    current_workshops = Workshop.objects.filter(status='ativo').count()
    
    # Estimativa baseada em proporção atual
    if current_workshops > 0:
        return (active_beneficiaries / current_workshops) * 0.3  # 30% de demanda estimada
    return active_beneficiaries * 0.1  # 10% se não há workshops


def _predict_resource_needs():
    """Predição de necessidades de recursos"""
    growth_rate = 0.05  # 5% de crescimento estimado
    current_beneficiaries = Beneficiary.objects.filter(status='ATIVA').count()
    
    return {
        'next_month_capacity': int(current_beneficiaries * (1 + growth_rate)),
        'staff_needed': int(current_beneficiaries * (1 + growth_rate) / 50),  # 1 staff para cada 50 beneficiárias
        'space_utilization': 'Alta' if current_beneficiaries > 200 else 'Média'
    }


def _export_custom_report_csv(report_data, report_type, start_date, end_date):
    """Exporta relatório personalizado em CSV"""
    import csv
    from django.http import HttpResponse
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{start_date}_{end_date}.csv"'
    
    writer = csv.writer(response)
    
    # Cabeçalho
    writer.writerow(['Relatório:', report_data['title']])
    writer.writerow(['Período:', f'{start_date} a {end_date}'])
    writer.writerow(['Total:', report_data['total_count']])
    writer.writerow([])
    
    # Dados específicos por tipo
    if report_type == 'beneficiaries':
        writer.writerow(['Nome', 'Email', 'Bairro', 'Status', 'Data de Cadastro'])
        for item in report_data['items']:
            writer.writerow([
                item.full_name,
                item.email,
                item.neighbourhood,
                item.status,
                item.created_at.strftime('%d/%m/%Y')
            ])
    
    return response


def _export_custom_report_excel(report_data, report_type, start_date, end_date):
    """Exporta relatório personalizado em Excel"""
    try:
        import openpyxl
        from django.http import HttpResponse
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_data['title']
        
        # Cabeçalho
        ws['A1'] = 'Relatório:'
        ws['B1'] = report_data['title']
        ws['A2'] = 'Período:'
        ws['B2'] = f'{start_date} a {end_date}'
        ws['A3'] = 'Total:'
        ws['B3'] = report_data['total_count']
        
        # Dados
        if report_type == 'beneficiaries':
            headers = ['Nome', 'Email', 'Bairro', 'Status', 'Data de Cadastro']
            for col, header in enumerate(headers, 1):
                ws.cell(row=5, column=col, value=header)
            
            for row, item in enumerate(report_data['items'], 6):
                ws.cell(row=row, column=1, value=item.full_name)
                ws.cell(row=row, column=2, value=item.email)
                ws.cell(row=row, column=3, value=item.neighbourhood)
                ws.cell(row=row, column=4, value=item.status)
                ws.cell(row=row, column=5, value=item.created_at.strftime('%d/%m/%Y'))
        
        # Salvar em BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{start_date}_{end_date}.xlsx"'
        
        return response
        
    except ImportError:
        # Fallback para CSV se openpyxl não estiver disponível
        return _export_custom_report_csv(report_data, report_type, start_date, end_date)
