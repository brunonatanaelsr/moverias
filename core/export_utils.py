# ==================================================
# CORE EXPORT UTILITIES
# Sistema centralizado de exportação para todos os módulos
# ==================================================

import csv
import io
from datetime import datetime
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
import logging

logger = logging.getLogger(__name__)

class ExportManager:
    """Gerenciador centralizado de exportação"""
    
    @staticmethod
    def export_to_csv(data, filename, headers=None):
        """
        Exporta dados para CSV com tratamento de erros
        """
        try:
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            
            writer = csv.writer(response)
            
            # Escrever cabeçalhos se fornecidos
            if headers:
                writer.writerow(headers)
            
            # Escrever dados
            for row in data:
                writer.writerow(row)
            
            return response
            
        except Exception as e:
            logger.error(f"Erro ao exportar CSV: {str(e)}")
            raise
    
    @staticmethod
    def export_to_excel(data, filename, headers=None, sheet_name="Dados"):
        """
        Exporta dados para Excel com formatação avançada
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Configurar cabeçalhos com formatação
            if headers:
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            # Inserir dados
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Salvar em BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            
            return response
            
        except ImportError:
            logger.warning("openpyxl não disponível, usando CSV como fallback")
            return ExportManager.export_to_csv(data, filename, headers)
        except Exception as e:
            logger.error(f"Erro ao exportar Excel: {str(e)}")
            raise
    
    @staticmethod
    def export_to_pdf(template_name, context, filename):
        """
        Exporta dados para PDF com formatação profissional
        """
        try:
            import weasyprint
            from django.conf import settings
            
            # Renderizar HTML
            html_string = render_to_string(template_name, context)
            
            # Configurações do PDF
            pdf_options = {
                'page-size': 'A4',
                'margin-top': '1cm',
                'margin-right': '1cm',
                'margin-bottom': '1cm',
                'margin-left': '1cm',
                'encoding': 'UTF-8',
            }
            
            # Gerar PDF
            pdf = weasyprint.HTML(
                string=html_string,
                base_url=settings.STATIC_URL
            ).write_pdf()
            
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
            
            return response
            
        except ImportError:
            logger.error("WeasyPrint não disponível para geração de PDF")
            raise ImportError("WeasyPrint é necessário para geração de PDF")
        except Exception as e:
            logger.error(f"Erro ao exportar PDF: {str(e)}")
            raise

class DataFormatter:
    """Formatador de dados para exportação"""
    
    @staticmethod
    def format_beneficiary_data(beneficiaries):
        """Formatar dados de beneficiárias para exportação"""
        headers = [
            'Nome Completo', 'Data de Nascimento', 'NIS', 'Telefone 1', 
            'Telefone 2', 'Bairro', 'Endereço', 'Referência', 
            'Status', 'Data de Cadastro'
        ]
        
        data = []
        for beneficiary in beneficiaries:
            data.append([
                beneficiary.full_name,
                beneficiary.dob.strftime('%d/%m/%Y') if beneficiary.dob else '',
                beneficiary.nis or '',
                beneficiary.phone_1 or '',
                beneficiary.phone_2 or '',
                beneficiary.neighbourhood or '',
                beneficiary.address or '',
                beneficiary.reference or '',
                beneficiary.get_status_display(),
                beneficiary.created_at.strftime('%d/%m/%Y %H:%M') if beneficiary.created_at else ''
            ])
        
        return data, headers
    
    @staticmethod
    def format_workshop_data(workshops):
        """Formatar dados de workshops para exportação"""
        headers = [
            'Nome', 'Tipo', 'Status', 'Data de Início', 'Data de Fim',
            'Max. Participantes', 'Participantes Atuais', 'Responsável', 'Data de Criação'
        ]
        
        data = []
        for workshop in workshops:
            data.append([
                workshop.name,
                workshop.get_workshop_type_display(),
                workshop.get_status_display(),
                workshop.start_date.strftime('%d/%m/%Y') if workshop.start_date else '',
                workshop.end_date.strftime('%d/%m/%Y') if workshop.end_date else '',
                workshop.max_participants,
                workshop.enrollments.count() if hasattr(workshop, 'enrollments') else 0,
                str(workshop.created_by) if hasattr(workshop, 'created_by') else '',
                workshop.created_at.strftime('%d/%m/%Y %H:%M') if workshop.created_at else ''
            ])
        
        return data, headers
    
    @staticmethod
    def format_task_data(tasks):
        """Formatar dados de tarefas para exportação"""
        headers = [
            'Título', 'Descrição', 'Status', 'Prioridade', 'Responsável',
            'Data de Criação', 'Data de Vencimento', 'Data de Conclusão'
        ]
        
        data = []
        for task in tasks:
            data.append([
                task.title,
                task.description[:100] + '...' if len(task.description) > 100 else task.description,
                task.get_status_display(),
                task.get_priority_display() if hasattr(task, 'get_priority_display') else '',
                str(task.assigned_to) if hasattr(task, 'assigned_to') else '',
                task.created_at.strftime('%d/%m/%Y %H:%M') if task.created_at else '',
                task.due_date.strftime('%d/%m/%Y') if hasattr(task, 'due_date') and task.due_date else '',
                task.completed_at.strftime('%d/%m/%Y %H:%M') if hasattr(task, 'completed_at') and task.completed_at else ''
            ])
        
        return data, headers
    
    @staticmethod
    def format_activity_data(activities):
        """Formatar dados de atividades para exportação"""
        headers = [
            'Nome', 'Tipo', 'Data', 'Beneficiária', 'Responsável',
            'Status', 'Observações', 'Data de Criação'
        ]
        
        data = []
        for activity in activities:
            data.append([
                activity.name,
                activity.get_activity_type_display() if hasattr(activity, 'get_activity_type_display') else '',
                activity.date.strftime('%d/%m/%Y') if hasattr(activity, 'date') and activity.date else '',
                str(activity.beneficiary) if hasattr(activity, 'beneficiary') else '',
                str(activity.responsible) if hasattr(activity, 'responsible') else '',
                activity.get_status_display() if hasattr(activity, 'get_status_display') else '',
                activity.notes[:100] + '...' if hasattr(activity, 'notes') and activity.notes and len(activity.notes) > 100 else (activity.notes if hasattr(activity, 'notes') else ''),
                activity.created_at.strftime('%d/%m/%Y %H:%M') if activity.created_at else ''
            ])
        
        return data, headers

def export_universal(request, model_class, formatter_method, filename_prefix, 
                    template_name=None, extra_context=None):
    """
    Função universal de exportação para qualquer modelo
    """
    export_format = request.GET.get('format', 'csv')
    
    # Obter dados
    queryset = model_class.objects.all().order_by('-created_at')
    
    # Aplicar filtros se fornecidos
    search = request.GET.get('search')
    if search:
        # Busca genérica (pode ser refinada por modelo)
        queryset = queryset.filter(name__icontains=search)
    
    # Formatar dados
    data, headers = formatter_method(queryset)
    
    # Gerar nome do arquivo
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{filename_prefix}_{timestamp}"
    
    try:
        if export_format == 'csv':
            return ExportManager.export_to_csv(data, filename, headers)
        
        elif export_format == 'excel':
            return ExportManager.export_to_excel(data, filename, headers, filename_prefix.title())
        
        elif export_format == 'pdf' and template_name:
            context = {
                'data': data,
                'headers': headers,
                'title': filename_prefix.replace('_', ' ').title(),
                'generated_at': timezone.now(),
                **(extra_context or {})
            }
            return ExportManager.export_to_pdf(template_name, context, filename)
        
        else:
            # Fallback para CSV
            return ExportManager.export_to_csv(data, filename, headers)
            
    except Exception as e:
        logger.error(f"Erro na exportação universal: {str(e)}")
        # Em caso de erro, retornar CSV básico
        return ExportManager.export_to_csv(data, filename, headers)
